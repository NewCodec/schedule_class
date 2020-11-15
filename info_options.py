from openpyxl import workbook
from openpyxl import load_workbook

from students import Student
from config import Config
from courses import Courses

def excle_load():
    xlsx = load_workbook("students_info.xlsx")
    info = xlsx["students_info"]
    schedule = xlsx["class_schedule"]
    config = xlsx["config"]
    return xlsx,info,schedule,config

#stdents is a list to contain all students
def load_students(sheet,students,config):
    flos = config.first_line_of_students
    # there is a number and name is stduents_info talbe
    while sheet.cell(flos,1).value!=None and sheet.cell(flos,2).value!=None :
        stu = Student()
        stu.number = sheet.cell(flos,1).value
        stu.name = sheet.cell(flos,2).value
        for i in range(7):
            stu.wish_time.append(sheet.cell(flos,3+i).value)
        stu.wish_courses_each_week = sheet.cell(flos,10).value
        stu.priority = sheet.cell(flos,11).value
        stu.info_print()
        students.append(stu)
        flos += 1

def init_courses(courses,config):
    for week in range(7):
        cours=[]
        for cur_num in range(config.max_courses_each_day) :
            cour = Courses(week,cur_num,config)
            cours.append(cour)
        courses.append(cours)

# find a course which can assinge a student
# not judge if the course is full.
def find_wish_courses(student,courses,config):
    for week in range(7):
        #match the course and wish
        wish = str(student.wish_time[week])
        #print(wish)
        match_cour = None
        if wish == 'None' :
            continue
        elif wish.isdigit():
            if student.wish_matched[week][int(wish)-1] == 0 :
                continue
            print("D:week=%d,wish=%d" % (week,int(wish)-1))
            match_cour = courses[week][int(wish)-1]
        else:
            for des in config.couse_description:
                index =int(des)-1
                print("S:week=%d,des=%d" % (week,index))
                if student.wish_matched[week][index] == 0 :
                    continue
                if wish == config.couse_description[des]:
                    match_cour = courses[week][index]
        if match_cour != None :
            return match_cour
    return match_cour
        
def deal_with_course_full(stu,courses,match_cour,config):
    if stu.priority != 1:
        stu.wish_matched[match_cour.week][match_cour.number] = 0
        return stu
    else :
        for sstu in match_cour.students:
            sstu.conflict = 1
            if sstu.priority != 1:
                delete_course_student(sstu,match_cour)
                sstu.conflict = 0
                break
        if sstu.conflict == 0:
            assigne_course_student(stu,match_cour,config)
            sstu.wish_matched[match_cour.week][match_cour.number] = 0
            print("reassigned %s" % sstu.name)
            return sstu
        else:
            sstu.conflict = 1
            print("%s confilicted",sstu.name)

def delete_course_student(student,course):
    print("delete %s from [%d][%d]" %(student.name,course.week,course.number))
    if course.students_count <=0 or student.scheduled_course_count<=0:
        return
    course.students.remove(student)
    course.students_count -= 1
    student.scheduled_courses.remove(course)
    student.scheduled_course_count -= 1 
    student.arranged_course -=1   

def assigne_course_student(student,course,config):
    course.students.append(student)
    course.students_count += 1
    student.scheduled_courses.append(course)
    student.scheduled_course_count += 1
    student.wish_matched[course.week][course.number] = 0
    student.arranged_course +=1
    if course.students_count == config.max_students_each_course :
        print("course[%d][%d] count=%d full" % (course.week,course.number,course.students_count))
        course.is_full = True

def inset_student_without_choice(stduent,courses,config):

    for num in range(config.max_courses_each_day):
        for week in range(7):       
            if courses[week][num].is_full == False and stduent.wish_matched[week][num]==1:
                assigne_course_student(stduent,courses[week][num],config)
                return
    print("ERROR:to much students!!!!")
    exit(1)

def schedule_courses(students,courses,config):

    for stu in students:
        print("========dealwith %s==========" % stu.name)
        sstu=stu
        while sstu.arranged_course < sstu.wish_courses_each_week :
            
            cour = find_wish_courses(sstu,courses,config)
            
            # 这个学生没有任何期望时间
            if cour == None or sstu.conflict == 1:
                print("None:%s",sstu.name)
                inset_student_without_choice(sstu,courses,config)
                continue;
            elif cour.is_full==True :
                sstu = deal_with_course_full(sstu,courses,cour,config)
                print("full %s "% sstu.name)
                continue
            else:
                assigne_course_student(sstu,cour,config)
            if sstu.arranged_course >= sstu.wish_courses_each_week and sstu != stu:
                sstu = stu

            #arranged_course+=1

        

def show_courses_to_file(sheet,courses,config):
    for week in range(7):
        for cur_num in range(config.max_courses_each_day) :
            names = ''
            for stu in courses[week][cur_num].students:
                names = names+' '+stu.name
            #print("[%d][%d]names=%s" % (week,cur_num,names))
            sheet.cell(row = cur_num+5,column=week+4).value = names   
        
    
