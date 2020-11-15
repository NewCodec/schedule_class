
from openpyxl import workbook
from openpyxl import load_workbook

class Config():
    def __init__(self,sheet):
        self.max_students_each_course = sheet.cell(23,3).value
        self.max_courses_each_day = 7
        self.max_priority = 1
        self.couse_description = {}
        for i in range(self.max_courses_each_day):
            self.couse_description[str(i+1)] = sheet.cell(4+i,4).value
        # line number of first students infor in stduents_info table
        self.first_line_of_students = 7



if __name__ == "__main__" :
    xlsx = load_workbook("students_info.xlsx")
    sheet = xlsx["config"]
    c = Config(sheet)
    print(sh)
    print(c.couse_description) 


